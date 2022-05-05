from asyncio.windows_events import NULL
from concurrent.futures import process
from pickle import NONE, TRUE
from tkinter import E
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl
import re
import pyperclip
import sys
import os

def search(s):
    excuteScript('selectCategory(0,0, true);')
    Search = getElement('ID', 'txtItemName')
    Search.clear()
    Search.send_keys(s)
    Search.send_keys('\n')
    time.sleep(1.5)

def changeWindow(n):
    browser.switch_to.window(browser.window_handles[n])

def login():
    if login_type != "stove":
        getElement('ID', "{}_login".format(login_type)).click()
        time.sleep(1)
        changeWindow(1)

    sid = {"stove"  : "user_id", 
           "naver"  : "id",
           "facebook" : "email",
           "twitter" : "username_or_email"}
    spw = {"stove" : "user_pwd",
           "naver" :  "pw",
           "facebook" : "pass",
           "twitter" : "password"}
    sclick = {"stove" : "#idLogin > div.row.grid.el-actions > button",
              "naver" : "#log\.login",
              "facebook" : "#loginbutton",
              "twitter" : "#allow"}

    bid = getElement('ID', sid[login_type])
    bid.click()
    pyperclip.copy(uid)
    bid.send_keys(Keys.CONTROL, 'v')
    time.sleep(0.5)

    bpw = getElement('ID', spw[login_type])
    bpw.click()
    pyperclip.copy(upw)
    bpw.send_keys(Keys.CONTROL, 'v')
    time.sleep(0.5)
    
    try:
        getElement('CSS_SELECTOR', sclick[login_type]).click()
        time.sleep(2)
        changeWindow(0)
    except:
        NONE

def showMessage():
    global progress
    progress += 1
    msg = "\r진행률 : %d%%"%(progress)
    print(msg, end='')

def getElement(by, val, isMany=False, _range=20, ignore=False):
    global bys, browser, err
    for i in range(_range):
        try:
            if isMany:
                return browser.find_elements(by=bys[by],value=val)
            else:                
                return browser.find_element(by=bys[by],value=val)
        except:
            if ignore:
                return NULL
            time.sleep(0.5)
    print("ERROR : ({}) when getElement {val}".format(err))
    time.sleep(5)
    sys.exit()

def excuteScript(val, _range=20):
    global browser, err
    for i in range(_range):
        try:
            browser.execute_script(val)
            return
        except:
            time.sleep(0.5)
    print("ERROR : ({}) when getScript {val}".format(err))
    time.sleep(5)
    sys.exit()
    

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
browser = webdriver.Chrome('chromedriver.exe', options=options)
browser.get('https://lostark.game.onstove.com/Market')

wb = openpyxl.load_workbook('base.xlsm', read_only=False, keep_vba=True)
ws = wb['거래소 최저가']

bys = {
    'CLASS_NAME' : By.CLASS_NAME,
    'CSS_SELECTOR' : By.CSS_SELECTOR,
    'ID' : By.ID,
}

os.system('cls')
print('\n','-'*43,'\n Lostark Craft Tool ver3.3.0.0 by 리퍼가신은신발은슬리퍼\n','-'*43,'\n')

login_type = wb['검색']['I5'].internal_value
uid = wb['검색']['I6'].internal_value
upw = wb['검색']['I7'].internal_value

# -------------------- Login --------------------

print("로봇 방지 등의 이유로 자동 로그인이 안됐다면 2분 내에 수동으로 로그인해주세요.")
print("로그인하면 자동으로 최저가를 탐색합니다.\n")

if login_type==None or uid==None or upw==None:
    browser.quit()
    print("로그인 정보 불러오기에 실패했습니다.")
    print("base.xlsm 파일에 로그인 정보를 잘 입력했는지 확인해주세요.")
    print("만약 Readme.txt를 읽지 않으셨다면 반드시 읽어주세요.")
    print("이 창은 10초 후 닫힙니다.")
    time.sleep(10)
    sys.exit()

try:
    login()
except Exception as e:
    print(e)
    print("로그인 과정에서 오류가 생겼습니다.")
    print("STOVE 계정이 아니라면 STOVE 계정으로 시도해보시기 바랍니다.")
    print("이 오류가 반복될 경우 제작자에게 문의해주세요.")
    print("이 창은 10초 후 닫힙니다.")
    time.sleep(10)
    sys.exit()

# -------------------- Logic --------------------

try:
    progress=-1
    err = 1
    elem = NULL
    elems = NULL

    elem = getElement('CSS_SELECTOR', '.main-category > li:nth-child(8) > a:nth-child(1)', _range=240)

    err+=1
    #progress:0
    showMessage()
    getElement('CSS_SELECTOR', '.main-category > li:nth-child(8) > a:nth-child(1)').click(); showMessage()
    getElement('CSS_SELECTOR', '.is-active > ul:nth-child(2) > li:nth-child(1) > a:nth-child(1)').click(); showMessage(); err+=1
    time.sleep(1.5); showMessage()
    getElement('CSS_SELECTOR', '#itemList > thead:nth-child(3) > tr:nth-child(1) > th:nth-child(1) > a:nth-child(1)').click(); showMessage(); err+=1

    cnt=0
    j=1
    #progress:4
    for k in range(1,5) :
        if k>1 :
            excuteScript('paging.page({});'.format(k))
        err+=1
        time.sleep(0.7); showMessage(); showMessage(); showMessage(); showMessage(); showMessage(); showMessage()
        prices = getElement('CLASS_NAME', "price", True); showMessage()
        names = getElement('CLASS_NAME', 'name', True); showMessage()
        cnt_name = 1; showMessage()
        for i in prices:
            cnt+=1
            if cnt==3 :
                cnt=0
                j+=1
                ws['A{}'.format(j)]=names[cnt_name].text
                try : 
                    ws['D{}'.format(j)] = re.sub(r'[^0-9]', '', (getElement('CSS_SELECTOR', '#tbodyItemList > tr:nth-child({}) > td:nth-child(1) > div:nth-child(1) > span:nth-child(3)'.format(cnt_name), ignore=True).text))
                except : None
                cnt_name+=1
                p=int(i.text.replace(',',""))
                ws['B{}'.format(j)] = p
        j+=1; err+=1
    getElement('CSS_SELECTOR', '.main-category > li:nth-child(6) > a:nth-child(1)').click()
    getElement('CSS_SELECTOR', '.is-active > ul:nth-child(2) > li:nth-child(1) > a:nth-child(1)').click(); err+=1
    time.sleep(1)
    getElement('CSS_SELECTOR', '#itemList > thead:nth-child(3) > tr:nth-child(1) > th:nth-child(1) > a:nth-child(1)').click(); err+=1

    cnt=0
    j=1
    #progress:40
    for k in range(1,7) :
        if k>1 :
            excuteScript('paging.page({});'.format(k))
        err+=1
        time.sleep(0.7); showMessage(); showMessage(); showMessage(); showMessage(); showMessage(); showMessage()
        prices = getElement('CLASS_NAME', "price", True)
        names = getElement('CLASS_NAME', 'name', True)
        cnt_name = 1
        for i in prices:
            cnt+=1
            if cnt==3 :
                cnt=0
                j+=1
                ws['F{}'.format(j)]=names[cnt_name].text
                cnt_name+=1
                p=int(i.text.replace(',',""))
                ws['G{}'.format(j)] = p
        j+=1; err+=1

    #progress:76
    search('하 융'); showMessage(); showMessage(); showMessage(); err+=1
    getElement('CSS_SELECTOR', '#itemList > thead:nth-child(3) > tr:nth-child(1) > th:nth-child(1) > a:nth-child(1)').click(); err+=1
    time.sleep(0.5)
    prices = getElement('CLASS_NAME', "price", True); err+=1
    ws['G61'] = int(prices[2].text.replace(',',"")); 
    ws['G62'] = int(prices[5].text.replace(',',"")); 
    ws['G63'] = int(prices[8].text.replace(',',"")); 

    #progress:83
    search('가루'); showMessage(); showMessage(); showMessage(); err+=1
    prices = getElement('CLASS_NAME', "price", True); err+=1
    ws['G67'] = int(prices[2].text.replace(',',""))
    
    #progress:86
    search('정식'); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.detail > div.grade > div > div.lui-select__title").click()
    time.sleep(0.5); showMessage(); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.detail > div.grade > div > div.lui-select__option > label:nth-child(4)").click()
    time.sleep(0.5); showMessage(); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.bt > button.button.button--deal-submit").click()
    time.sleep(0.5); showMessage(); err+=1
    prices = getElement('CLASS_NAME', "price", True)
    names = getElement('CLASS_NAME', 'name', True); err+=1
    cnt_name = 0
    cnt=0
    j=69
    for i in prices:
        cnt+=1
        if cnt==3 :
            cnt=0
            cnt_name+=1
            try : 
                err+=1
                tmp = getElement('CSS_SELECTOR', '#tbodyItemList > tr:nth-child({}) > td:nth-child(1) > div:nth-child(1) > span:nth-child(3)'.format(cnt_name), ignore=True).text; 
                ws['F{}'.format(j)]=names[cnt_name].text; showMessage(); showMessage()
                p=int(i.text.replace(',',""))
                ws['G{}'.format(j)] = p
                j+=1        
            except : None

    search('꼬치'); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.detail > div.grade > div > div.lui-select__title").click()
    time.sleep(0.5); showMessage(); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.detail > div.grade > div > div.lui-select__option > label:nth-child(3)").click()
    time.sleep(0.5); showMessage(); err+=1
    getElement('CSS_SELECTOR', "#lostark-wrapper > div > main > div > div.deal > div.deal-contents > form > fieldset > div > div.bt > button.button.button--deal-submit").click()
    time.sleep(0.5); showMessage(); err+=1
    prices = getElement('CLASS_NAME', "price", True)
    names = getElement('CLASS_NAME', 'name', True); err+=1
    cnt_name = 0
    cnt=0
    j=73
    for i in prices:
        cnt+=1
        if cnt==3 :
            cnt=0
            cnt_name+=1
            try : 
                err+=1
                tmp = getElement('CSS_SELECTOR', '#tbodyItemList > tr:nth-child({}) > td:nth-child(1) > div:nth-child(1) > span:nth-child(3)'.format(cnt_name), ignore=True).text; 
                ws['F{}'.format(j)]=names[cnt_name].text; showMessage(); showMessage()
                p=int(i.text.replace(',',""))
                ws['G{}'.format(j)] = p
                j+=1        
            except : None

    err = 9999
    t = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
    wb.save('{}.xlsm'.format(t)); showMessage()
    progress=99; showMessage()

    print('\n\n최저가 데이터 동기화 완료. \n{}.xlsm 파일에서 확인하세요.\n이 창은 5초 후 자동으로 꺼집니다.'.format(t))
    time.sleep(5)
    browser.quit()
except Exception as e:
    print("ERROR Progress:{}".format(progress))
    print("동일한 에러가 반복될 경우 제작자에게 문의해주세요.")
    print("5초 후 종료됩니다.")
    time.sleep(5)
    